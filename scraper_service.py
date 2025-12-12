# -*- coding: utf-8 -*-
import re, time, json, os
import requests
from bs4 import BeautifulSoup, NavigableString
from openpyxl import Workbook
from typing import List, Dict, Tuple

HEADERS = {"User-Agent": "Mozilla/5.0"}

# -------------------- Utils --------------------
def tclean(s: str) -> str:
    return re.sub(r"\s+", " ", s or "").strip()

def get_soup(url: str) -> BeautifulSoup:
    r = requests.get(url, headers=HEADERS, timeout=30)
    r.raise_for_status()
    return BeautifulSoup(r.text, "lxml")

def heading_level(tag):
    return int(tag.name[1]) if tag and tag.name and tag.name.startswith("h") and tag.name[1].isdigit() else 99

# -------------------- Infobox & sections --------------------
def parse_infobox_map(soup: BeautifulSoup):
    data = {}
    for row in soup.select(".portable-infobox .pi-data"):
        lab = row.select_one(".pi-data-label")
        val = row.select_one(".pi-data-value")
        if lab and val:
            data[tclean(lab.get_text())] = tclean(val.get_text(" ", strip=True))
    return data

def extract_section_node(soup: BeautifulSoup, titles):
    patt = re.compile("|".join([re.escape(t) for t in titles]), flags=re.I)
    for h in soup.select("h2,h3,h4"):
        if patt.search(h.get_text(" ", strip=True)):
            return h
    return None

def extract_section_text(soup: BeautifulSoup, titles):
    start = extract_section_node(soup, titles)
    if not start:
        return ""
    out = []
    for sib in start.find_all_next():
        if sib.name and sib.name.startswith("h"):
            break
        if sib.name in ("p", "li", "blockquote"):
            txt = tclean(sib.get_text(" ", strip=True))
            if txt:
                out.append(txt)
    return "\n".join(out)

# -------------------- Overall (JSON-LD abstract) --------------------
def extract_intro_text(soup: BeautifulSoup) -> str:
    content = soup.select_one(".mw-parser-output")
    if not content:
        return ""
    
    out = []
    # Iterate direct children to grab text before the first section or TOC
    for node in content.children:
        # Ignore NavigableStrings that are just whitespace
        if isinstance(node, NavigableString) and not node.strip():
            continue
            
        # Skip Infobox (aside) or Table of Contents (already handled below but good to be safe)
        if node.name == "aside" or (node.name == "div" and "portable-infobox" in node.get("class", [])):
            continue

        if node.name == "p":
            # Remove nested infoboxes if they exist inside the paragraph
            for bad in node.select("aside, .portable-infobox"):
                bad.decompose()

            txt = tclean(node.get_text(" ", strip=True))
            if txt:
                out.append(txt)
        
        # Stop at the first real section header or Table of Contents
        elif node.name == "h2":
            break
        elif node.name == "div" and node.get("id") == "toc":
            break
            
    return "\n".join(out)

# -------------------- Quotes (Voice Lines + Other Quotes) --------------------
def extract_quotes(soup: BeautifulSoup) -> str:
    def gather_by_id_or_text(anchor_id, text_keyword):
        span = soup.find("span", {"id": anchor_id})
        if not span:
            for h in soup.select("h2,h3,h4"):
                if text_keyword.lower() in h.get_text(" ", strip=True).lower():
                    span = h.find("span", class_="mw-headline") or h
                    break
        if not span:
            return []
        htag = span if span.name in ("h2","h3","h4") else span.find_parent(["h2","h3","h4"])
        if not htag:
            return []
        ul = htag.find_next("ul")  # không cần sibling trực tiếp
        if not ul:
            return []
        lines = []
        for li in ul.find_all("li"):
            txt = tclean(li.get_text())
            if txt:
                lines.append(txt)
        return lines

    quotes = []
    quotes += gather_by_id_or_text("Voice_Lines", "Voice Lines")
    quotes += gather_by_id_or_text("Other_Quotes", "Other Quotes")
    # khử trùng lặp
    seen, out = set(), []
    for q in quotes:
        if q not in seen:
            out.append(q); seen.add(q)
    return "\n".join(out)

# -------------------- Roles (Primary / Secondary) --------------------
ROLE_SEP_RE = re.compile(r"\s*[,/•|;]\s*")

def extract_roles_from_infobox(inf: dict) -> dict:
    out = {"primary": "", "secondary": ""}

    # 1) Có trường rõ ràng
    for k, v in inf.items():
        lk = k.lower()
        if "primary role" in lk or lk.strip() == "primary":
            out["primary"] = v
        elif "secondary role" in lk or lk.strip() == "secondary":
            out["secondary"] = v

    if out["primary"] or out["secondary"]:
        return out

    # 2) Chỉ có Role/Roles hoặc Classification
    role_key = next((k for k in inf.keys() if k.lower() in ("role","roles")), None)
    if not role_key:
        role_key = next((k for k in inf.keys() if "classification" in k.lower()), None)

    if role_key:
        text = inf[role_key]

        # Pattern “Primary: X … Secondary: Y …”
        m = re.search(r"(?i)primary\s*[:\-]\s*([^;|•\n]+)", text)
        if m:
            out["primary"] = m.group(1).strip()
        m = re.search(r"(?i)secondary\s*[:\-]\s*([^;|•\n]+)", text)
        if m:
            out["secondary"] = m.group(1).strip()
        if out["primary"] or out["secondary"]:
            return out

        # Nếu là danh sách vai trò
        parts = [p for p in ROLE_SEP_RE.split(text) if p]
        if parts:
            out["primary"] = parts[0]
            if len(parts) > 1:
                out["secondary"] = ", ".join(parts[1:])
            return out

    return out

# -------------------- Skills --------------------
def extract_skills_from_table(soup: BeautifulSoup) -> list:
    skills_heading = extract_section_node(soup, ["Skills"])
    if not skills_heading:
        return []
    skills_table = None
    for tb in skills_heading.find_all_next("table"):
        prev_h = tb.find_previous(lambda tag: getattr(tag, "name", "").startswith("h"))
        if prev_h and prev_h is not skills_heading:
            break
        headers = [tclean(th.get_text(" ", strip=True)).lower() for th in tb.select("tr th")]
        if headers and any("name" in h for h in headers) and any("description" in h for h in headers):
            skills_table = tb
            break
    if not skills_table:
        return []

    header_cells = [tclean(th.get_text(" ", strip=True)).lower()
                    for th in skills_table.select("tr")[0].find_all(["th","td"])]
    def idx_of(key, default=None):
        for i,h in enumerate(header_cells):
            if key in h:
                return i
        return default
    name_idx = idx_of("name")
    desc_idx = idx_of("description")
    unlock_idx = idx_of("unlock")
    icon_idx = idx_of("icon") # New: Look for icon column

    if name_idx is None or desc_idx is None:
        return []

    out = []
    for tr in skills_table.select("tr")[1:]:
        tds = tr.find_all(["td","th"])
        if len(tds) <= max(name_idx, desc_idx):
            continue
        name = tclean(tds[name_idx].get_text(" ", strip=True))
        desc = tclean(tds[desc_idx].get_text(" ", strip=True))
        unlock = tclean(tds[unlock_idx].get_text(" ", strip=True)) if unlock_idx is not None and len(tds)>unlock_idx else ""
        
        # Extract Icon
        icon = ""
        if icon_idx is not None and len(tds) > icon_idx:
            img = tds[icon_idx].find("img")
            if img:
                icon = img.get("data-src") or img.get("src") or ""

        if name or desc or unlock:
            out.append({"Unlock Level": unlock, "Skill Name": name, "Type": "", "Description": desc, "Icon": icon})
    return out

def extract_skills_boxes(soup: BeautifulSoup) -> list:
    out = []
    for box in soup.select("div.skillbox"):
        header = box.select_one(".skillbox-header")
        if not header:
            continue
        name = tclean(header.get_text(" ", strip=True))
        typ = ""
        small = header.find("small")
        if small:
            typ = tclean(small.get_text(" ", strip=True))
        
        # Extract Icon
        icon = ""
        img_div = box.select_one(".skillbox-image")
        if img_div:
            img = img_div.find("img")
            if img:
                icon = img.get("data-src") or img.get("src") or ""

        desc_parts = []
        for d in box.select(".skillbox-description, p, li"):
            txt = tclean(d.get_text(" ", strip=True))
            if txt:
                desc_parts.append(txt)
        # bỏ trùng dòng
        seen, dedup = set(), []
        for line in desc_parts:
            if line not in seen:
                dedup.append(line); seen.add(line)
        desc = "\n".join(dedup)
        if name or desc:
            out.append({"Unlock Level": "", "Skill Name": name, "Type": typ, "Description": desc, "Icon": icon})
    return out

def extract_skills(soup: BeautifulSoup) -> list:
    skills = extract_skills_from_table(soup)
    if skills:
        return skills
    return extract_skills_boxes(soup)

# -------------------- Engraving Abilities (sheet riêng) --------------------
def extract_engraving_rows(soup: BeautifulSoup, hero_name: str) -> list:
    root = extract_section_node(soup, ["Engraving Abilities","Engraving"])
    if not root:
        return []
    # ưu tiên bảng nếu có
    table = None
    for tb in root.find_all_next("table"):
        prev_h = tb.find_previous(lambda tag: getattr(tag, "name","").startswith("h"))
        if prev_h and prev_h is not root:
            break
        headers = [tclean(th.get_text(" ", strip=True)).lower() for th in tb.select("tr th")]
        if headers and (any("unlock" in h for h in headers) or any("level" in h for h in headers)) and any("description" in h for h in headers):
            table = tb
            break
    rows = []
    if table:
        header_cells = [tclean(th.get_text(" ", strip=True)).lower()
                        for th in table.select("tr")[0].find_all(["th","td"])]
        def idx_of(key):
            for i,h in enumerate(header_cells):
                if key in h:
                    return i
            return None
        name_idx = idx_of("name")
        desc_idx = idx_of("description")
        unlock_idx = idx_of("unlock") or idx_of("level")
        icon_idx = idx_of("icon") # Icon column

        for tr in table.select("tr")[1:]:
            tds = tr.find_all(["td","th"])
            if not tds: continue
            name = tclean(tds[name_idx].get_text(" ", strip=True)) if name_idx is not None and len(tds)>name_idx else ""
            desc = tclean(tds[desc_idx].get_text(" ", strip=True)) if desc_idx is not None and len(tds)>desc_idx else ""
            unlock = tclean(tds[unlock_idx].get_text(" ", strip=True)) if unlock_idx is not None and len(tds)>unlock_idx else ""
            
            icon = ""
            if icon_idx is not None and len(tds) > icon_idx:
                img = tds[icon_idx].find("img")
                if img:
                    icon = img.get("data-src") or img.get("src") or ""

            if name or desc or unlock:
                rows.append({"Hero": hero_name, "Skill Name": name, "Unlock Level": unlock, "Description": desc, "Icon": icon})
        return rows

    # fallback: tách theo mốc E30/E60/E80 trong đoạn văn
    text = extract_section_text(soup, ["Engraving Abilities","Engraving"])
    parts = re.split(r"(?i)\b(E\s*30|E\s*60|E\s*80|\[30\]|\[60\]|\[80\])", text)
    for i in range(1, len(parts), 2):
        mark = parts[i].strip("[] ").upper().replace(" ", "")
        desc = tclean(parts[i+1]) if i+1 < len(parts) else ""
        rows.append({"Hero": hero_name, "Skill Name": "", "Unlock Level": mark, "Description": desc, "Icon": ""})
    return rows

# -------------------- Furniture Set Bonuses (sheet riêng) --------------------
def extract_furniture_rows(soup: BeautifulSoup, hero_name: str) -> list:
    root = extract_section_node(soup, ["Furniture Set Bonuses","Furniture"])
    if not root:
        return []
    rows = []
    
    # The name is usually in the next H3 or H4
    # Example: H2 Furniture -> H3 Name -> UL Description
    name_header = None
    for h in root.find_all_next(["h3", "h4"]):
        # Stop if we hit another H2 (next main section)
        if h.name == "h2":
            break
        # Stop if we hit the limit of section (rough heuristic)
        if h.find_previous(lambda t: t is root):
            name_header = h
            break
            
    name = tclean(name_header.get_text(" ", strip=True)) if name_header else ""
    
    # Try to find an image near the header
    icon = ""
    if name_header:
        # Check specific containers often used in Fandom
        # Case 1: <figure> before or after
        target_img = None
        
        # Look backwards for figure or image
        prev = name_header.find_previous(["figure", "div"])
        if prev and (prev.get("class") and any("tright" in c for c in prev.get("class",[])) or prev.name=="figure"):
             target_img = prev.find("img")

        # If not found, look inside the header (sometimes icon is inline)
        if not target_img:
            target_img = name_header.find("img")

        # If not found, look forward
        if not target_img:
             nxt = name_header.find_next(["figure", "div"])
             if nxt and (nxt.get("class") and any("tright" in c for c in nxt.get("class",[])) or nxt.name=="figure"):
                 target_img = nxt.find("img")

        if target_img:
            icon = target_img.get("data-src") or target_img.get("src") or ""

    ul = (name_header.find_next("ul") if name_header else root.find_next("ul"))
    desc_lines = []
    if ul:
        for li in ul.find_all("li"):
            txt = tclean(li.get_text(" ", strip=True))
            if txt: desc_lines.append(txt)
    desc = "\n".join(desc_lines)
    
    if name or desc:
        rows.append({"Hero": hero_name, "Name": name or "Furniture", "Description": desc, "Icon": icon})
    return rows

# -------------------- Signature Item (sheet riêng – Hero | Description) --------------------
def find_sig_heading(soup):
    span = soup.find("span", id=re.compile(r"^signature[\s_]*item$", re.I))
    if span:
        h = span.find_parent(["h2","h3"])
        if h: return h
    for h in soup.select("h2,h3"):
        if re.search(r"\bsignature\s+item\b", h.get_text(" ", strip=True), re.I):
            return h
    return None

def extract_signature_item_desc(soup: BeautifulSoup) -> str:
    root = find_sig_heading(soup)
    if not root:
        return ""
    base = heading_level(root)
    lines = []

    for node in root.find_all_next():
        if node is root:
            continue
        if node.name and node.name.startswith("h") and heading_level(node) <= base:
            break

        if isinstance(node, NavigableString):
            continue

        if node.name in ("h3", "h4"):
            title = tclean(node.get_text(" ", strip=True))
            if re.match(r"(?i)^(item|skill)\s*:", title):
                lines.append(title)
            continue

        if node.name in ("p","blockquote"):
            txt = tclean(node.get_text(" ", strip=True))
            if txt:
                lines.append(txt)
        elif node.name == "ul":
            for li in node.find_all("li"):
                txt = tclean(li.get_text(" ", strip=True))
                if txt:
                    lines.append(txt)

    seen, out = set(), []
    for line in lines:
        if line and line not in seen:
            out.append(line); seen.add(line)
    return "\n".join(out).strip()

def extract_signature_rows(soup: BeautifulSoup, hero_name: str) -> list:
    desc = extract_signature_item_desc(soup)
    return [{"Hero": hero_name, "Description": desc}] if desc else []

# -------------------- Scrape 1 hero --------------------
# -------------------- Scrape 1 hero --------------------
def scrape_page(url: str):
    try:
        soup = get_soup(url)
        title_elem = soup.select_one("#firstHeading")
        title = title_elem.get_text(strip=True) if title_elem else "Unknown"

        # Infobox + icon
        inf_map = parse_infobox_map(soup)
        roles = extract_roles_from_infobox(inf_map)
        icon = ""
        infobox = soup.select_one(".portable-infobox")
        if infobox:
            img = infobox.select_one("img")
            if img and img.has_attr("src"):
                icon = img["src"]

        intro = extract_intro_text(soup)
        personality = extract_section_text(soup, ["Personality"])
        

        # Prepare dict with EN/VN suffixes
        # Map original keys to _EN keys
        hero_row = {
            "Icon": icon,
            "Name": title,
            "Faction": inf_map.get("Faction",""),
            "Type":    inf_map.get("Type",""),
            "Class":   inf_map.get("Class",""),
            "Rarity":  inf_map.get("Rarity",""),
            "Role":    inf_map.get("Role",""),
            "Primary Role": roles.get("primary",""),
            "Secondary Role": roles.get("secondary",""),
            "Overall_EN": f"{intro}\n\n{personality}".strip(),
            "Overall_VN": "",
            "Personality_EN": personality,
            "Personality_VN": "",
            "Background_EN":  extract_section_text(soup, ["Background","Story"]),
            "Background_VN": "",
            "Quotes_EN":      extract_quotes(soup),
            "Quotes_VN":      "",
            "Trivia_EN":      extract_section_text(soup, ["Trivia"]),
            "Trivia_VN":      "",
            "URL": url,
        }

        # Process Sub-lists to have _EN / _VN
        skills = extract_skills(soup)
        for s in skills:
            s["Hero"] = title
            s.setdefault("Type","")
            # Rename Description -> Description_EN
            desc = s.pop("Description", "")
            s["Description_EN"] = desc
            s["Description_VN"] = ""

        engraving_rows = extract_engraving_rows(soup, title)
        for r in engraving_rows:
            desc = r.pop("Description", "")
            r["Description_EN"] = desc
            r["Description_VN"] = ""

        signature_rows = extract_signature_rows(soup, title)
        for r in signature_rows:
            desc = r.pop("Description", "")
            r["Description_EN"] = desc
            r["Description_VN"] = ""

        furniture_rows = extract_furniture_rows(soup, title)
        for r in furniture_rows:
            desc = r.pop("Description", "")
            r["Description_EN"] = desc
            r["Description_VN"] = ""
        
        # --- Translation ---
        # Removed per user request


        return hero_row, skills, engraving_rows, signature_rows, furniture_rows
    except Exception as e:
        print(f"Error scraping {url}: {e}")
        import traceback
        traceback.print_exc()
        return None, [], [], [], []

from urllib.parse import urljoin

# ... existing imports ...

# -------------------- Faction Page Scanning --------------------
def find_heroes_section(soup):
    # Priority: <span id="Heroes">
    span = soup.find("span", id=re.compile(r"^Heroes$", re.I))
    if span:
        h = span.find_parent(["h2", "h3"])
        if h: return h
    # Fallback: search h2/h3 text
    for h in soup.select("h2,h3"):
        if re.search(r"\bHeroes\b", h.get_text(" ", strip=True), re.I):
            return h
    return None

def extract_hero_links(faction_url: str) -> List[str]:
    try:
        soup = get_soup(faction_url)
        root = find_heroes_section(soup)
        if not root:
            return []

        seen, out = set(), []
        # Iterate siblings until next header
        for node in root.find_all_next():
            if node is root:
                continue
            if getattr(node, "name", "") in ("h2", "h3"):
                break
            
            for a in node.find_all("a", href=True):
                name = tclean(a.get_text(" ", strip=True))
                # Basic filter: name shouldn't be empty, link should be relative aka internal wiki link
                if not name or not a["href"].startswith("/wiki/"):
                    continue
                
                href = urljoin(faction_url, a["href"])
                if href in seen:
                    continue
                
                # Check duplication and potentially invalid links (like Category/File)
                if ":" in a["href"].replace("/wiki/", ""): 
                    # Skip special pages mostly
                    continue

                # Filter out specific non-hero utility links
                ignored_paths = ["/wiki/Heroes", "/wiki/Rarity", "/wiki/Class", "/wiki/Type", "/wiki/Faction", "/wiki/Union"]
                if any(a["href"].startswith(p) for p in ignored_paths):
                    continue

                seen.add(href)
                out.append(href)
        return out
    except Exception as e:
        print(f"Error scanning faction: {e}")
        return []

# -------------------- Write Excel --------------------
def create_excel(heroes, skills, engraving_rows, signature_rows, furniture_rows, output_path: str):
    wb = Workbook()

    # Heroes
    hero_cols = ["ID","Icon","Name","Faction","Type","Class","Rarity","Role",
                 "Primary Role","Secondary Role",
                 "Overall","Personality","Background","Quotes","Trivia","URL"]
    ws1 = wb.active; ws1.title = "Heroes"
    ws1.append(hero_cols)
    for h in heroes:
        ws1.append([h.get(c,"") for c in hero_cols])

    # Skills
    ws2 = wb.create_sheet("Skills")
    ws2.append(["Hero", "Hero ID", "Unlock Level","Skill Name","Type","Description"])
    for s in skills:
        ws2.append([s.get("Hero",""), s.get("Hero ID",""), s.get("Unlock Level",""), s.get("Skill Name",""), s.get("Type",""), s.get("Description","")])

    # Engraving Abilities
    ws3 = wb.create_sheet("Engraving Abilities")
    ws3.append(["Hero", "Hero ID", "Skill Name","Unlock Level","Description"])
    for r in engraving_rows:
        ws3.append([r.get("Hero",""), r.get("Hero ID",""), r.get("Skill Name",""), r.get("Unlock Level",""), r.get("Description","")])

    # Signature Item
    ws4 = wb.create_sheet("Signature Item")
    ws4.append(["Hero", "Hero ID", "Description"])
    for r in signature_rows:
        ws4.append([r.get("Hero",""), r.get("Hero ID",""), r.get("Description","")])

    # Furniture Set Bonuses
    ws5 = wb.create_sheet("Furniture Set Bonuses")
    ws5.append(["Hero", "Hero ID", "Name","Description"])
    for r in furniture_rows:
        ws5.append([r.get("Hero",""), r.get("Hero ID",""), r.get("Name",""), r.get("Description","")])

    wb.save(output_path)
    return output_path
